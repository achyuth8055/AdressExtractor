import os
import uuid
import time
import re
import json
import threading
from datetime import datetime, timedelta

from flask import Flask, request, render_template, send_file, jsonify, session
from werkzeug.utils import secure_filename

import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "ccta-extractor-secret-key-2026")
app.config["UPLOAD_FOLDER"] = os.path.join(os.path.dirname(__file__), "uploads")
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024

os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)

jobs = {}


def create_driver():
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")

    chrome_bin = os.environ.get("CHROME_BIN")
    chromedriver_path = os.environ.get("CHROMEDRIVER_PATH")

    if chrome_bin:
        options.binary_location = chrome_bin
    if chromedriver_path:
        service = Service(chromedriver_path)
        return webdriver.Chrome(service=service, options=options)
    return webdriver.Chrome(options=options)


def search_pin(driver, pin):
    p1, p2, p3, p4, p5 = pin[0:2], pin[2:4], pin[4:7], pin[7:10], pin[10:14]
    max_retries = 3
    for attempt in range(1, max_retries + 1):
        try:
            driver.get("https://www.cookcountytreasurer.com/setsearchparameters.aspx")
            wait = WebDriverWait(driver, 20)
            wait.until(EC.presence_of_element_located(
                (By.ID, "ContentPlaceHolder1_ASPxPanel1_SearchByPIN1_txtPIN1")
            ))
            driver.execute_script("""
                document.getElementById('ContentPlaceHolder1_ASPxPanel1_SearchByPIN1_txtPIN1').value = arguments[0];
                document.getElementById('ContentPlaceHolder1_ASPxPanel1_SearchByPIN1_txtPIN2').value = arguments[1];
                document.getElementById('ContentPlaceHolder1_ASPxPanel1_SearchByPIN1_txtPIN3').value = arguments[2];
                document.getElementById('ContentPlaceHolder1_ASPxPanel1_SearchByPIN1_txtPIN4').value = arguments[3];
                document.getElementById('ContentPlaceHolder1_ASPxPanel1_SearchByPIN1_txtPIN5').value = arguments[4];
            """, p1, p2, p3, p4, p5)
            driver.execute_script(
                "document.getElementById('ContentPlaceHolder1_ASPxPanel1_SearchByPIN1_cmdContinue').click();"
            )
            wait.until(EC.url_contains("yourpropertytaxoverviewresults"))
            time.sleep(3)
            body_text = driver.find_element(By.TAG_NAME, "body").text
            result = parse_mailing(body_text)
            if result:
                return result
            if attempt < max_retries:
                time.sleep(2)
        except Exception:
            if attempt < max_retries:
                time.sleep(3)
    return None


def parse_mailing(body_text):
    lines = body_text.split("\n")
    mailing_idx = None
    for i, line in enumerate(lines):
        if "Mailing Information:" in line:
            mailing_idx = i
            break
    if mailing_idx is None:
        return None
    mailing_lines = []
    for j in range(mailing_idx + 1, min(mailing_idx + 10, len(lines))):
        stripped = lines[j].strip()
        if not stripped:
            continue
        if stripped.lower().startswith("update your"):
            break
        mailing_lines.append(stripped)
        if len(mailing_lines) >= 3:
            break
    if len(mailing_lines) < 3:
        return None
    result = {"name": mailing_lines[0], "address": mailing_lines[1],
              "city": "", "state": "", "zipcode": ""}
    match = re.match(r"^(.+?),\s+(\w{2})\s+(\S+)$", mailing_lines[2])
    if match:
        result["city"] = match.group(1).strip()
        result["state"] = match.group(2).strip()
        result["zipcode"] = match.group(3).strip()
    else:
        result["city"] = mailing_lines[2]
    return result


def process_job(job_id, file_path):
    job = jobs[job_id]
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        pin_col, name_col, addr_col, city_col, state_col, zip_col = 1, 2, 3, 4, 5, 6

        pins = []
        for row_num in range(2, ws.max_row + 1):
            pin_val = ws.cell(row=row_num, column=pin_col).value
            if pin_val:
                pin_str = str(pin_val).strip().replace("-", "").replace(" ", "")
                if pin_str.endswith(".0"):
                    pin_str = pin_str[:-2]
                pin_str = pin_str.zfill(14)
                pins.append((row_num, pin_str))

        total = len(pins)
        job["total"] = total
        if total == 0:
            job["status"] = "error"
            job["log"].append("No PINs found in the file.")
            job["done"] = True
            return

        job["log"].append(f"Found {total} PINs. Starting browser...")
        driver = create_driver()
        success = 0
        failed = 0

        try:
            for idx, (row_num, pin) in enumerate(pins, 1):
                job["progress"] = idx
                job["current_pin"] = pin
                job["log"].append(f"[{idx}/{total}] Looking up PIN: {pin}")

                if len(pin) != 14 or not pin.isdigit():
                    job["log"].append(f"  ✗ Invalid PIN format: {pin}")
                    failed += 1
                    continue

                result = search_pin(driver, pin)
                if result:
                    ws.cell(row=row_num, column=name_col, value=result["name"])
                    ws.cell(row=row_num, column=addr_col, value=result["address"])
                    ws.cell(row=row_num, column=city_col, value=result["city"])
                    ws.cell(row=row_num, column=state_col, value=result["state"])
                    ws.cell(row=row_num, column=zip_col, value=result["zipcode"])
                    job["log"].append(f"  ✓ {result['name']}, {result['address']}, {result['city']}, {result['state']} {result['zipcode']}")
                    success += 1
                else:
                    job["log"].append(f"  ✗ Failed to retrieve after 3 attempts")
                    failed += 1

                if idx < total:
                    time.sleep(1)
        finally:
            driver.quit()

        wb.save(file_path)
        job["status"] = "complete"
        job["log"].append(f"\nDone! {success} succeeded, {failed} failed.")
        job["success"] = success
        job["failed"] = failed

    except Exception as e:
        job["status"] = "error"
        job["log"].append(f"Error: {str(e)}")
    finally:
        job["done"] = True


def cleanup_old_files():
    folder = app.config["UPLOAD_FOLDER"]
    cutoff = time.time() - 3600
    for fname in os.listdir(folder):
        fpath = os.path.join(folder, fname)
        if os.path.isfile(fpath) and os.path.getmtime(fpath) < cutoff:
            os.remove(fpath)


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/upload", methods=["POST"])
def upload():
    cleanup_old_files()

    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No file selected"}), 400

    if not file.filename.endswith(".xlsx"):
        return jsonify({"error": "Only .xlsx files are accepted"}), 400

    job_id = str(uuid.uuid4())[:8]
    filename = f"{job_id}_{secure_filename(file.filename)}"
    file_path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
    file.save(file_path)

    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        if "DT_PIN" not in headers:
            os.remove(file_path)
            return jsonify({"error": "Excel file must have a 'DT_PIN' column in the first row."}), 400
        pin_count = sum(1 for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value)
        wb.close()
    except Exception as e:
        os.remove(file_path)
        return jsonify({"error": f"Could not read Excel file: {str(e)}"}), 400

    jobs[job_id] = {
        "status": "running",
        "progress": 0,
        "total": pin_count,
        "current_pin": "",
        "log": [],
        "file_path": file_path,
        "filename": file.filename,
        "done": False,
        "success": 0,
        "failed": 0,
    }

    thread = threading.Thread(target=process_job, args=(job_id, file_path), daemon=True)
    thread.start()

    return jsonify({"job_id": job_id, "total": pin_count})


@app.route("/status/<job_id>")
def status(job_id):
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    return jsonify({
        "status": job["status"],
        "progress": job["progress"],
        "total": job["total"],
        "current_pin": job["current_pin"],
        "log": job["log"],
        "done": job["done"],
        "success": job.get("success", 0),
        "failed": job.get("failed", 0),
    })


@app.route("/download/<job_id>")
def download(job_id):
    job = jobs.get(job_id)
    if not job:
        return "Job not found", 404
    if not job["done"] or job["status"] != "complete":
        return "File not ready", 400
    return send_file(
        job["file_path"],
        as_attachment=True,
        download_name=job["filename"],
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port, debug=False)
